import time
import requests
import sys
import urllib.parse
import openpyxl
import argparse
from veracode_api_signing.plugin_requests import RequestsAuthPluginVeracodeHMAC
from veracode_api_signing.credentials import get_credentials

headers = {"User-Agent": "Veracode eLearning Script"}
api_base = "https://api.veracode.{intance}"
auth = RequestsAuthPluginVeracodeHMAC()

learners_cache = dict()

THROTTLE_LIMIT = 10

def update_api_base():
    global api_base
    api_key_id, _ = get_credentials()
    if api_key_id.startswith("vera01"):
        api_base = api_base.replace("{intance}", "eu", 1)
    else:
        api_base = api_base.replace("{intance}", "com", 1)

def has_more_pages(json_response):
    if not "page" in json_response:
        return False
    page_node = json_response["page"]
    current_page = page_node["number"]
    total_pages = page_node["totalPages"]
    return current_page < (total_pages-1)

def get_courses(page=0):
    global api_base
    global auth
    courses_endpoint = f"{api_base}/elearning/v1/courses?size=500&page={page}"

    response = requests.get(courses_endpoint, auth=auth, headers=headers)

    if response and response.ok:
        courses = response.json()["_embedded"]["courses"]
        print(f"Successfully fetched courses page {page}. Found {len(courses)} courses")
        if has_more_pages(response.json()):
            return courses + get_courses(page+1)
        else:    
            return courses
    else:
        print(f"ERROR: unable to fetch courses {response.status_code}")
        if response and response.json():
            print(f"-- {response.json()}")
        print()
        sys.exit(-1)

def get_learner(learner_url):
    if learner_url in learners_cache:
        return learners_cache[learner_url]
    
    global THROTTLE_LIMIT
    global api_base
    global auth

    response = requests.get(learner_url, auth=auth, headers=headers)

    learner_descriptor = ""
    if response and response.ok:
        learner = response.json()
        if learner:
            learner_descriptor = f"{learner['firstName']} {learner['lastName']} ({learner['email']})"

    learners_cache[learner_url] = learner_descriptor
    return learner_descriptor
        
    
def get_report_cards_for_course(course, page=0, attempt=0):
    global THROTTLE_LIMIT
    global api_base
    global auth

    report_cards_endpoint = f"{api_base}/elearning/v1/reportcards?course_id={urllib.parse.quote_plus(course['courseId'])}&size=500&page={page}"

    response = requests.get(report_cards_endpoint, auth=auth, headers=headers)

    if response and response.ok:
        report_cards = response.json()["_embedded"]["reportcards"]
        print(f"Successfully fetched report cards page {page} for course {course['courseId']}. Found {len(report_cards)} report cards")
        if has_more_pages(response.json()):
            return report_cards + get_report_cards_for_course(course, page+1, attempt)
        else:    
            return report_cards
    elif response and response.status_code == 429:
        print(f"WARNING: hit 429 status, waiting 1 minute")
        time.sleep(60)
        if attempt < THROTTLE_LIMIT:
            return get_report_cards_for_course(course, page, attempt+1)
        else:
            print(f"ERROR: exceeded 429 limit ({THROTTLE_LIMIT}). Exiting")
            sys.exit(-1)
    else:
        print(f"ERROR: unable to fetch report cards page {page} for course {course['courseId']} - {response.status_code}")
        if response and response.json():
            print(f"-- {response.json()}")

        #TODO: check courseId C++MEM issue
        if course["courseId"] == "C++MEM":
            return []
        
        print()
        sys.exit(-1)


def save_report(file_name):
    all_courses = get_courses()
    all_courses_info = []
    for course in all_courses:
        all_courses_info.append({"course": course, "report_cards": get_report_cards_for_course(course)})

    try:
        excel_file = openpyxl.Workbook()
        excel_sheet = excel_file.active

        try:
            current_row = 1
            excel_sheet.cell(row = current_row, column = 1).value="Course"
            excel_sheet.cell(row = current_row, column = 2).value="Learner"
            excel_sheet.cell(row = current_row, column = 3).value="Course Status"
            
            for course_information in all_courses_info:
                report_cards = course_information["report_cards"]
                course_name = course_information["course"]["name"]

                if report_cards:
                    for report_card in report_cards:
                        current_row += 1
                        learner_id = report_card["_links"]["learner"]["href"]
                        excel_sheet.cell(row = current_row, column = 1).value=course_name
                        excel_sheet.cell(row = current_row, column = 2).value=get_learner(learner_id)
                        excel_sheet.cell(row = current_row, column = 3).value=report_card["courseStatus"]
                else:
                    current_row += 1
                    excel_sheet.cell(row = current_row, column = 1).value=course_name
                    excel_sheet.cell(row = current_row, column = 2).value="NO LEARNERS"
                    excel_sheet.cell(row = current_row, column = 3).value=""

        finally:
            excel_file.save(filename=file_name)
            print(f"Successfully saved report to {file_name}")
    except Exception as e:
        print(f"ERROR: unable to save report to {file_name}")
        print(e)
        sys.exit(-1)

def main():
    parser = argparse.ArgumentParser(
        description="This script creates a Veracode eLearning report."
    )

    parser.add_argument(
        "-f",
        "--file_name",
        help="Name of the xlsx file to save.",
        required=True
    )

    args = parser.parse_args()

    file_name = args.file_name.strip().lower()
    if not (".xlsx" == file_name[-5:]):
        print("ERROR: -f/--file_name extension must be .xlsx")
        sys.exit(-1)

    update_api_base()
    save_report(file_name)

if __name__ == "__main__":
    main()
